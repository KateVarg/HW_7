import os
import zipfile
from conftest import RESOURCES_DIR, TEST_PDF, TEST_XLSX, TEST_CSV
import csv
from openpyxl import load_workbook
from pypdf import PdfReader


def test_csv():
    with zipfile.ZipFile(os.path.join(RESOURCES_DIR, "new_archive.zip")) as zip_file:
        with zip_file.open(TEST_CSV) as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            eight_row = csvreader[8]

            assert eight_row[3] == 'Female'
            assert eight_row[5] == '27'


def test_xlsx():
    with zipfile.ZipFile(os.path.join(RESOURCES_DIR, "new_archive.zip")) as zip_file:
        with zip_file.open(TEST_XLSX) as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            print(sheet.cell(row=3, column=2).value)

            assert sheet.cell(row=3, column=2).value == 'Mara'


def test_pdf():
    with zipfile.ZipFile(os.path.join(RESOURCES_DIR, "new_archive.zip")) as zip_file:
        with zip_file.open(TEST_PDF) as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[1]
            text = page.extract_text()

            assert 'Please don\'t break this trust' in text
